from openpyxl.styles import PatternFill, Font, Alignment, numbers, Border, Side
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from os.path import isfile
import pandas as pd

class Objection:
  def __init__(self, framePath, fdate):
    self.__dataframe = pd.read_csv(framePath)
    self.__dfPath = framePath
    self.__filter_day = fdate
    self.__destPath = ""
    self.format_accounting = u'_($* #,##0.00_);[Red]_($* (#,##0.00);_($* -_0_0_);_(@'

  def Full_Form(self):
    # Anio en formato numerico
    format_year = str(datetime.now().year)

    #Get current month name
    months = {"January": "Enero", "February": "Febrero", "March": "Marzo", "April": "Abril", "May": "Mayo", "June": "Junio", "July": "Julio", "August": "Agosto", "September": "Septiembre", "October": "Octubre", "November": "Noviembre", "December": "Diciembre"}
    eng_month = str(datetime.now().strftime('%B'))
    format_month = months[eng_month]
    
    #self.__destPath = "/Users/goldenfield/Google Drive File Stream/Unidades Compartidas/Contabilidad/BLACK PEARL TEA/CONTABILIDAD/{year}/Reportes Wan/12. {month} 20/CONCENTRADO DE VENTAS.xlsx".format(year=format_year, month=format_month)
    self.__destPath = "/Users/goldenfield/Documents/Backups/Contabilidad/Cedulas/CONCENTRADO DE VENTAS.xlsx".format(year=format_year, month=format_month)
    #self.__destPath = "G:/Unidades Compartidas/Contabilidad/BLACK PEARL TEA/CONTABILIDAD/{year}/Reportes Wan/11. {month} 20/CONCENTRADO DE VENTAS.xlsx".format(year=format_year, month=format_month)

    days = {"Monday": "LUNES", "Tuesday": "MARTES", "Wednesday": "MIERCOLES", "Thursday": "JUEVES", "Friday": "VIERNES", "Saturday": "SABADO", "Sunday": "DOMINGO"}
    months = {"January": "ENERO","February": "FEBRERO","March": "MARZO","April": "ABRIL","May": "MAYO","June": "JUNIO","July": "JULIO","August": "AGOSTO","September": "SEPTIEMBRE","October": "OCTUBRE","November": "NOVIEMBRE","December": "DICIEMBRE"}

    fday = days[str(self.__filter_day.strftime("%A"))]
    fmonth = months[str(self.__filter_day.strftime("%B"))]
    fday_num = str(self.__filter_day.strftime("%d"))

    complete_date = str(self.__filter_day)


    if not isfile(self.__destPath):
      Book = Workbook()
      Sheet = Book.active
      Sheet.title = "EFECTIVOS"
      last_column = -1
    else:
      Book = load_workbook(self.__destPath)
      Sheet = Book["EFECTIVOS"]
      last_column = Sheet.max_column

    icolumn = last_column + 2
    columns = ["", complete_date]

    # Create headers in excel file
    for col, val in enumerate(columns):
      destiny = icolumn + col
      Sheet.cell(row=3, column=destiny).font = Font(size=13, bold=True, color="00000000")
      Sheet.cell(row=3, column=destiny).fill = PatternFill(fill_type="solid", start_color="00C0C0C0", end_color="00C0C0C0")
      Sheet.cell(row=3, column=destiny).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
      Sheet.cell(row=3, column=destiny).border = Border(left=Side(border_style='medium', color='00000000'), right=Side(border_style='medium', color='00000000'), top=Side(border_style='medium', color='00000000'), bottom=Side(border_style='medium',color='00000000'))
      Sheet.cell(row=3, column=destiny, value=val)
      Sheet.column_dimensions[get_column_letter(destiny)].width = 15.5

    Sheet.column_dimensions[get_column_letter(1)].width = 30
    Sheet.column_dimensions[get_column_letter(11)].width = 10

    Sheet.row_dimensions[3].height = 40

    data = self.__dataframe[["Subtotal"]]
    stores = self.__dataframe["store_name"].values.tolist()

    frow = 4
    for name in stores:
      row = data[self.__dataframe["store_name"] == name]
      Sheet.cell(row=frow, column=icolumn).value = name
      Sheet.cell(row=frow, column=icolumn).font = Font(bold=True, color="00000000")
      amount_total = 0.0
      for idx, amount in enumerate(row.values.tolist()[0]):
          dest = icolumn + idx
          amount_total = amount_total + float(amount)
          Sheet.cell(row=frow, column=dest+1).value = float(amount)
          Sheet.cell(row=frow, column=dest+1).number_format= self.format_accounting
      frow = frow + 1
    
    
    Book.save(self.__destPath)